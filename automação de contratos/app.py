from openpyxl import load_workbook
from docx import Document
from datetime import datetime


# Obter dados da planilha "fornecedores";
data_vendors = load_workbook('./fornecedores.xlsx')
sheet_data_vendors = data_vendors['Sheet1']

for line in sheet_data_vendors.iter_rows(min_row=2, values_only=True):
    ...

# Obter dados do documento "contrato.txt";
# Adicionar dados coletados em posições específicas do contrato;
# Salvar novo contrato em uma pasta específica;

